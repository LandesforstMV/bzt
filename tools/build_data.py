#!/usr/bin/env python3
"""
Erzeugt aus der BZT-Zieltabelle (xlsx) die Datendatei fuer die Filter-App.

    python3 tools/build_data.py

Eingabe : data/B5_BZT_Zieltabelle_01.wide.All_Appedit.xlsx
          data/Baumarten.Code.DSW.xlsx  (Klartext- und Artnamen)
Ausgabe : data/bzt_data.js       (window.BZT_DATA = {...})
          data/bzt_data.json     (identischer Inhalt, z.B. fuer R/Python)
          assets/DATEILISTE.md   (Liste aller erwarteten Bilddateien)

Wenn die Zieltabelle aktualisiert wird, einfach dieses Skript erneut laufen
lassen - die HTML-/JS-Dateien der App muessen dafuer nicht angefasst werden.
"""
from __future__ import annotations

import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("Bitte zuerst 'pip install openpyxl' ausfuehren.")

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "B5_BZT_Zieltabelle_01.wide.All_Appedit.xlsx"
XLSX_BA = ROOT / "data" / "Baumarten.Code.DSW.xlsx"
OUT_JS = ROOT / "data" / "bzt_data.js"
OUT_JSON = ROOT / "data" / "bzt_data.json"
OUT_LIST = ROOT / "assets" / "DATEILISTE.md"

# ---------------------------------------------------------------------------
# BESCHRIFTUNGEN - hier bei Bedarf anpassen
# ---------------------------------------------------------------------------

# Klimastufen: Schluessel = Wert in der Spalte "Klimastufe"
KLIMASTUFEN = OrderedDict([
    ("Tf, Tlf, Tlm", {
        "id": "Tf",
        "kurz": "Tf",
        "name": "feuchtes Tieflandklima",
        "beschreibung": "einschl. Tlf und Tlm",
    }),
    ("Tm", {
        "id": "Tm",
        "kurz": "Tm",
        "name": "mäßig trockenes Tieflandklima",
        "beschreibung": "",
    }),
    ("Tt", {
        "id": "Tt",
        "kurz": "Tt",
        "name": "trockenes Tieflandklima",
        "beschreibung": "",
    }),
])

# Naehrstoffstufen (2. Buchstabe im Standortgruppen-Kuerzel)
NAEHRSTUFEN = OrderedDict([
    ("R", "reich"),
    ("K", "kräftig"),
    ("M", "mittel"),
    ("Z", "ziemlich arm"),
    ("A", "arm"),
])

# Reihenfolge der Standort-Gruppenueberschriften in der Standort-Spalte.
# Von den terrestrischen Standorten zu den nassesten. Die Werte muessen
# denen der Spalte "Standortbeschreibung" entsprechen.
GRUPPEN_REIHENFOLGE = [
    "terrestrische Standorte",
    "wechselfrische Staustandorte",
    "wechselnasse/wechselfeuchte Standorte",
    "mineralische Naßstandorte",
    "organische Naßstandorte",
    "Überwässerungs-Standorte",
]

# Praefix im Standortgruppen-Kuerzel -> Wasserhaushalt (Kurzform).
# Die Langform wird aus der Spalte "Standortbeschreibung" uebernommen.
PRAEFIXE = OrderedDict([
    ("Ue", "Überwässerung"),
    ("O", "organisch nass"),
    ("N", "mineralisch nass"),
    ("T", "wechselfrisch (Stau)"),
    ("", "terrestrisch"),
])

# Suffixe im Standortgruppen-Kuerzel
SUFFIXE = {
    "w": "wechselfeucht / wechselnass",
    "(ue)": "zeitweise überflutet",
    "ue": "zeitweise überflutet",
    "g": "grundfeucht",
    "(g)": "grundfeucht",
    "+": "Plus-Standort",
}

# Feuchtestufen des Standorts-Piktogramms, woertlich aus dem Erlass S. 8/9:
#   trocken (T..3, T..+3), maessig frisch (T..2, T..2g, T..+2),
#   frisch (T..1, T..+1, T..1w, W..2, N..3, N..+3),
#   feucht (Ue..2, N..2, N..2w, O..4, O..4w, O..4ue),
#   nass (Ue..1, N..1, N..1w, O..3, O..3ue),
#   sumpfig (Ue..0, N..0, O..2, O..1)
# Schluessel: Praefix des Kuerzels, dann Wasserstufe -> Feuchtestufe.
# Terrestrische Standorte tragen kein Praefix, wechselfrische Staustandorte
# das Praefix T; beide folgen derselben Regel.
FEUCHTE_STUFEN = ["trocken", "mäßig frisch", "frisch", "feucht", "nass", "sumpfig"]
FEUCHTE_REGELN = {
    "":   {3: "trocken", 2: "mäßig frisch", 1: "frisch"},
    "T":  {3: "trocken", 2: "mäßig frisch", 1: "frisch"},
    "N":  {3: "frisch", 2: "feucht", 1: "nass", 0: "sumpfig"},
    "O":  {4: "feucht", 3: "nass", 2: "sumpfig", 1: "sumpfig"},
    "Ue": {2: "feucht", 1: "nass", 0: "sumpfig"},
}

# Farben der Standortskarte (Erlass S. 66: Feuchtestufe = Farbton,
# Naehrkraft = Saettigung). Reihenfolge der Naehrstufen: R K M Z A.
STANDORT_FARBEN = {
    "":   ["#2f8f21", "#49b02f", "#69c94b", "#93dc76", "#bfeaa9"],  # terrestrisch
    "T":  ["#2f8f21", "#49b02f", "#69c94b", "#93dc76", "#bfeaa9"],  # Staustandorte
    "N":  ["#9e2049", "#c72f61", "#e34c7c", "#ef7fa1", "#f8b3c7"],  # mineralisch nass
    "O":  ["#1a2fc0", "#2c47e0", "#5570ee", "#8b9df4", "#bfc8f9"],  # organisch nass
    "Ue": ["#6b21b8", "#8b34d6", "#a95ee6", "#c491f0", "#dcbdf7"],  # Ueberwaesserung
}
# Achtung, zwei Reihenfolgen: die Farbpalette oben ist von reich nach arm
# sortiert, die Spalten des Standorts-Piktogramms laufen umgekehrt von arm
# nach reich (Erlass, Abschnitt 3.5). Deshalb zwei getrennte Indizes.
NAEHR_REIHE = ["R", "K", "M", "Z", "A"]      # Farbpalette: reich -> arm
PIKTO_SPALTEN = ["A", "Z", "M", "K", "R"]    # Piktogramm: arm -> reich

# ---------------------------------------------------------------------------


def lies_baumarten() -> dict:
    """Klartext- und wissenschaftliche Namen aus der DSW-Kuerzelliste."""
    if not XLSX_BA.exists():
        sys.exit(f"Baumartenliste nicht gefunden: {XLSX_BA}")
    wb = openpyxl.load_workbook(XLSX_BA, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.values)
    header = [str(h) if h is not None else "" for h in rows[0]]
    out = {}
    for r in rows[1:]:
        d = dict(zip(header, r))
        code = (str(d.get("WERT_KF") or "")).strip()
        if not code:
            continue
        out[code] = {
            "name": (str(d.get("WERT_LF") or code)).strip(),
            "latVoll": (str(d.get("BEMERK") or "")).strip(),
            "lat": kurzer_artname(d.get("BEMERK")),
            "gruppe": "NB" if str(d.get("Laub_Nadel") or "").startswith("Nadel") else "LB",
            "obergruppe": (str(d.get("TransferName") or "")).strip(),
        }
    return out


def kurzer_artname(bemerk) -> str:
    """'Fagus sylvatica L.' -> 'Fagus sylvatica', 'Larix x eurolepis HENRY'
    -> 'Larix x eurolepis'. Autorenkuerzel und Klammerzusaetze entfallen."""
    if not bemerk:
        return ""
    text = re.sub(r"\([^)]*\)", " ", str(bemerk))
    teile = [t for t in text.replace(",", " ").split() if t]
    if not teile:
        return ""
    name = [teile[0]]
    rest = teile[1:]
    if rest and rest[0].lower() in ("x", "×"):
        name += rest[:2]
    elif rest:
        name.append(rest[0])
    return " ".join(name)


# ---------------------------------------------------------------------------

UMLAUTE = {
    "ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss",
    "Ä": "Ae", "Ö": "Oe", "Ü": "Ue",
}


def slug(text: str) -> str:
    """Dateiname-taugliches Kuerzel (ASCII, keine Sonderzeichen)."""
    out = "".join(UMLAUTE.get(ch, ch) for ch in str(text))
    out = out.replace("+", "plus")
    out = re.sub(r"[()\s]+", "", out)
    out = re.sub(r"[^A-Za-z0-9_-]", "_", out)
    return out


STGR_RE = re.compile(r"^(Ü|O|N|T)?([RKMZA])(\d)(.*)$")


def parse_stgr(code: str) -> dict:
    """Zerlegt ein Standortgruppen-Kuerzel wie 'NR2w' oder 'M2+'."""
    m = STGR_RE.match(code)
    if not m:
        return {"praefix": "", "naehr": "", "stufe": "", "suffix": ""}
    praefix, naehr, stufe, suffix = m.groups()
    praefix = {"Ü": "Ue"}.get(praefix or "", praefix or "")
    return {
        "praefix": praefix,
        "naehr": naehr,
        "stufe": stufe,
        "suffix": (suffix or "").strip(),
    }


def split_arten(value) -> list:
    if not value:
        return []
    return [s.strip() for s in str(value).split(",") if s.strip()]


def num(value):
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    return int(f) if f == int(f) else round(f, 2)


def main() -> int:
    if not XLSX.exists():
        sys.exit(f"Zieltabelle nicht gefunden: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.values)
    header = [str(h) if h is not None else "" for h in rows[0]]
    records = [dict(zip(header, r)) for r in rows[1:] if r[0] is not None]

    # --- Klimastufen -------------------------------------------------------
    klima_keys = [k for k in KLIMASTUFEN if any(r["Klimastufe"] == k for r in records)]
    unbekannt = {r["Klimastufe"] for r in records} - set(KLIMASTUFEN)
    if unbekannt:
        sys.exit(f"Unbekannte Klimastufe(n) in der Tabelle: {sorted(unbekannt)}")
    klima = []
    for key in klima_keys:
        cfg = dict(KLIMASTUFEN[key])
        cfg["quellwert"] = key
        cfg["slug"] = slug(cfg["id"])
        klima.append(cfg)
    klima_idx = {k["quellwert"]: i for i, k in enumerate(klima)}

    # --- Standortgruppen ---------------------------------------------------
    standorte = []
    gesehen = set()
    for r in records:
        code = r["STGR_BZT"]
        if code in gesehen:
            continue
        gesehen.add(code)
        parts = parse_stgr(code)
        suffix_key = parts["suffix"].replace("ü", "ue").lower()
        stufe_idx = NAEHR_REIHE.index(parts["naehr"]) if parts["naehr"] in NAEHR_REIHE else 2
        palette = STANDORT_FARBEN.get(parts["praefix"], STANDORT_FARBEN[""])
        feuchte = FEUCHTE_REGELN.get(parts["praefix"], {}).get(int(parts["stufe"]))
        if not feuchte:
            sys.exit(f"Keine Feuchtestufe fuer '{code}' – FEUCHTE_REGELN pruefen.")
        standorte.append({
            "farbe": palette[stufe_idx],
            "feuchte": feuchte,
            "naehrIdx": stufe_idx,
            "piktoSpalte": PIKTO_SPALTEN.index(parts["naehr"]),
            "feuchteIdx": FEUCHTE_STUFEN.index(feuchte),
            "id": code,
            "slug": slug(code),
            "kartiercode": r["STGR_Kart"],
            "gruppe": r["Standortbeschreibung"],
            "wasserhaushalt": PRAEFIXE.get(parts["praefix"], parts["praefix"]),
            "naehrstoff": NAEHRSTUFEN.get(parts["naehr"], parts["naehr"]),
            "naehrCode": parts["naehr"],
            "stufe": parts["stufe"],
            "zusatz": SUFFIXE.get(suffix_key, parts["suffix"]),
            "zusatzLang": r["Standortszusatz"],
        })

    # Gruppen in der oben festgelegten Reihenfolge; innerhalb einer Gruppe
    # bleibt die Reihenfolge der Zieltabelle erhalten (reich -> arm).
    unbekannte_gruppen = {s_["gruppe"] for s_ in standorte} - set(GRUPPEN_REIHENFOLGE)
    if unbekannte_gruppen:
        sys.exit(f"Standortgruppe ohne Platz in GRUPPEN_REIHENFOLGE: "
                 f"{sorted(unbekannte_gruppen)}")
    standorte.sort(key=lambda s_: GRUPPEN_REIHENFOLGE.index(s_["gruppe"]))
    stgr_idx = {s_["id"]: i for i, s_ in enumerate(standorte)}

    # --- Bestandeszieltypen ------------------------------------------------
    bzt = []
    bzt_idx = {}
    for r in records:
        key = r["BZT_Nr"]
        if key in bzt_idx:
            continue
        gruppen = []
        for i in (1, 2, 3):
            arten = split_arten(r[f"Baumart_{i}"])
            if not arten:
                continue
            gruppen.append({
                "rang": i,
                "label": str(r[f"Anteil_BA{i}"] or "").strip(),
                "min": num(r[f"Min_Anteil_BA{i}"]),
                "max": num(r[f"Max_Anteil_BA{i}"]),
                "arten": arten,
            })
        bzt_idx[key] = len(bzt)
        bzt.append({
            "id": key,
            "nr": num(r["Nr"]),
            "slug": slug(key),
            "typ": r["BZT.Typ"],
            "name": r["BZT_Bezeichnung"],
            "lhAnteil": num(r["LH_Anteil"]),
            "baAnzahl": num(r["BA_Anzahl"]),
            "seite": num(r["Seiten"]),
            "gruppen": gruppen,
        })
    bzt.sort(key=lambda b: b["nr"])
    bzt_idx = {b["id"]: i for i, b in enumerate(bzt)}

    # Konsistenzpruefung: Zusammensetzung haengt nur am BZT
    signatur = {}
    for r in records:
        sig = json.dumps(
            [[r[f"Baumart_{i}"], r[f"Anteil_BA{i}"], r[f"Min_Anteil_BA{i}"],
              r[f"Max_Anteil_BA{i}"]] for i in (1, 2, 3)],
            ensure_ascii=False,
        )
        if signatur.setdefault(r["BZT_Nr"], sig) != sig:
            sys.exit(
                f"{r['BZT_Nr']}: Baumartenzusammensetzung ist nicht eindeutig - "
                "die App setzt eine je BZT konstante Zusammensetzung voraus."
            )

    # --- Baumarten ---------------------------------------------------------
    verwendet = []
    for b in bzt:
        for g in b["gruppen"]:
            verwendet.extend(g["arten"])
    katalog = lies_baumarten()
    fehlend = sorted(set(verwendet) - set(katalog))
    if fehlend:
        print(f"Hinweis: Kuerzel fehlen in {XLSX_BA.name}: {fehlend}", file=sys.stderr)
    baumarten = []
    for code in sorted(set(verwendet)):
        e = katalog.get(code, {"name": code, "lat": "", "latVoll": "",
                               "gruppe": "LB", "obergruppe": ""})
        baumarten.append({"code": code, "slug": slug(code), **e})
    baumarten.sort(key=lambda a: (a["gruppe"] == "NB", a["name"]))

    # --- Kombinationen -----------------------------------------------------
    kombis = sorted({
        (klima_idx[r["Klimastufe"]], stgr_idx[r["STGR_BZT"]], bzt_idx[r["BZT_Nr"]])
        for r in records
    })

    for name, eintraege in [("Standortgruppen", standorte), ("BZT", bzt),
                            ("Baumarten", baumarten), ("Klimastufen", klima)]:
        gesehen = {}
        for e in eintraege:
            key = e["slug"]
            if key in gesehen:
                sys.exit(f"{name}: '{e.get('id', e.get('code'))}' und "
                         f"'{gesehen[key]}' ergeben denselben Dateinamen '{key}.png'. "
                         "Bitte slug() in diesem Skript anpassen.")
            gesehen[key] = e.get("id", e.get("code"))

    daten = {
        "meta": {
            "zeilen": len(records),
            "kombinationen": len(kombis),
        },
        "klimastufen": klima,
        "standorte": standorte,
        "bzt": bzt,
        "baumarten": baumarten,
        "kombis": [list(k) for k in kombis],
    }

    OUT_JSON.write_text(
        json.dumps(daten, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    OUT_JS.write_text(
        "/* Automatisch erzeugt von tools/build_data.py - nicht von Hand aendern. */\n"
        "window.BZT_DATA = "
        + json.dumps(daten, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )

    # --- Dateiliste fuer die Bilder ---------------------------------------
    endungen = [".png", ".jpg", ".jpeg", ".webp", ".svg"]

    def vorhanden(ordner: str, name: str):
        for e in endungen:
            pfad = ROOT / "assets" / ordner / f"{name}{e}"
            if pfad.exists():
                return f"{name}{e}"
        return None

    lines = [
        "# Kachelbilder",
        "",
        "Automatisch erzeugt von `tools/build_data.py`.",
        "",
        "Die App sucht zu jedem Eintrag der Reihe nach `.png`, `.jpg`, `.jpeg`,",
        "`.webp`, `.svg`. Ein selbst abgelegtes **PNG gewinnt** damit immer gegen",
        "eine mitgelieferte JPG- oder SVG-Datei gleichen Namens - die",
        "Originaldatei muss nicht geloescht werden. Fehlt jede Datei, zeichnet",
        "die App einen Platzhalter.",
        "",
        "Die mit **[ok]** markierten Bilder sind vorhanden: Karten und",
        "Bestandesbilder aus `tools/extract_images.py`, Blattzeichnungen aus",
        "`tools/make_baumartenblaetter.py`, Standorts-Piktogramme aus",
        "`tools/make_standortpiktogramme.py`.",
        "",
        "Empfohlen: quadratisch oder 4:3 (Klimastufe/Standort/Baumart) bzw. sehr",
        "breit (BZT-Bestandesbild), mindestens 480 px Kantenlaenge.",
        "",
    ]
    for titel, ordner, eintraege in [
        ("Klimastufen (Karten aus Klimastufen_BZT.pdf)", "klimastufe",
         [(k["slug"], f"{k['kurz']} - {k['name']}") for k in klima]),
        ("Standortgruppen (Piktogramm Naehrkraft x Feuchtestufe)", "standort",
         [(s_["slug"], f"{s_['id']} - {s_['gruppe']}") for s_ in standorte]),
        ("Bestandeszieltypen (Bestandesbilder aus BZT_Erlass.pdf)", "bzt",
         [(b["slug"], f"{b['typ']} - {b['name']} (Erlass S. {b['seite']})") for b in bzt]),
        ("Baumarten (schematische Blattzeichnungen)", "baumart",
         [(a["slug"], f"{a['code']} - {a['name']}") for a in baumarten]),
    ]:
        da = sum(1 for name, _ in eintraege if vorhanden(ordner, name))
        lines.append(f"## {titel}")
        lines.append("")
        lines.append(f"`assets/{ordner}/` - {da} von {len(eintraege)} vorhanden")
        lines.append("")
        for name, beschr in eintraege:
            datei = vorhanden(ordner, name)
            marke = f"**[ok]** `{datei}`" if datei else f"`{name}.png` (fehlt)"
            lines.append(f"- {marke} - {beschr}")
        lines.append("")

    lines += [
        "## Logos und Favicon",
        "",
        "Diese Dateien sind optional; fehlende werden stillschweigend",
        "ausgeblendet.",
        "",
        "- `assets/favicon.ico` - Symbol im Browser-Tab",
        "- `assets/logo.png` - kleines Logo links oben neben dem Titel",
        "- `assets/logo_lf.png` - Logo oben mittig",
        "- `assets/logo1.png` bis `assets/logo4.png` - Logoleiste unten rechts",
        "",
    ]
    OUT_LIST.write_text("\n".join(lines), encoding="utf-8")

    print(f"{len(records)} Zeilen gelesen")
    print(f"  {len(klima)} Klimastufen, {len(standorte)} Standortgruppen, "
          f"{len(bzt)} BZT, {len(baumarten)} Baumarten, {len(kombis)} Kombinationen")
    print(f"geschrieben: {OUT_JS.relative_to(ROOT)}, {OUT_JSON.relative_to(ROOT)}, "
          f"{OUT_LIST.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
